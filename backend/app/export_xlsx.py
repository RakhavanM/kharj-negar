from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from .jalali import MONTH_NAMES, to_jalali

CATEGORY_LABELS = {
    "daily": "خرج روزمره",
    "installment": "قسط",
    "rent": "اجاره",
    "car": "ماشین",
    "home": "وسایل خانه",
    "debt": "قرض",
    "pet": "پت",
    "miscellaneous": "متفرقه",
}
PERSON_LABELS = {"ramin": "رامین", "mana": "مانا"}

DARK_GREEN = "163300"
WISE_GREEN = "9FE870"
MINT = "E2F6D5"
WHITE = "FFFFFF"


def _excel_datetime(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _style_header(row) -> None:
    fill = PatternFill("solid", fgColor=DARK_GREEN)
    font = Font(name="Aptos", bold=True, color=WHITE)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_title(cell, fill_color: str = MINT) -> None:
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(name="Aptos Display", bold=True, size=16, color=DARK_GREEN)
    cell.alignment = Alignment(horizontal="right", vertical="center")


def _add_table(ws, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _set_widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def create_expenses_workbook(
    expenses: Iterable,
    users: Iterable | None = None,
    household_name: str = "خرج‌نگار",
    category_labels: dict[str, str] | None = None,
) -> bytes:
    expenses = list(expenses)
    users = list(users or [])
    category_labels = {**CATEGORY_LABELS, **(category_labels or {})}
    workbook = Workbook()
    expenses_sheet = workbook.active
    expenses_sheet.title = "هزینه‌ها"
    summary_sheet = workbook.create_sheet("خلاصه ماهانه")

    expense_headers = [
        "شناسه", "مبلغ (تومان)", "مبلغ (هزار تومان)", "فرد خرج‌کننده",
        "دسته‌بندی", "تاریخ شمسی", "تاریخ میلادی", "توضیحات", "ثبت‌کننده",
        "زمان ثبت (UTC)", "آخرین ویرایش (UTC)",
    ]
    expenses_sheet.append([f"خروجی کامل هزینه‌های {household_name}"])
    expenses_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(expense_headers))
    _style_title(expenses_sheet["A1"])
    expenses_sheet.row_dimensions[1].height = 28
    expenses_sheet.append(expense_headers)
    _style_header(expenses_sheet[2])
    expenses_sheet.freeze_panes = "A3"
    expenses_sheet.auto_filter.ref = f"A2:K{max(2, len(expenses) + 2)}"

    for expense in expenses:
        expenses_sheet.append([
            expense.id,
            expense.amount_toman,
            expense.amount_toman // 1000,
            PERSON_LABELS.get(expense.person, expense.person),
            category_labels.get(expense.category, expense.category),
            to_jalali(expense.expense_date),
            expense.expense_date,
            expense.note or "",
            getattr(expense.created_by, "username", "") if getattr(expense, "created_by", None) else "",
            _excel_datetime(expense.created_at),
            _excel_datetime(expense.updated_at),
        ])

    for row in expenses_sheet.iter_rows(min_row=3, max_row=expenses_sheet.max_row):
        row[1].number_format = '#,##0 "تومان"'
        row[2].number_format = '#,##0'
        row[6].number_format = "yyyy-mm-dd"
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="right")
    if expenses_sheet.max_row >= 3:
        _add_table(expenses_sheet, f"A2:K{expenses_sheet.max_row}", "ExpensesExport")
    _set_widths(expenses_sheet, {"A": 10, "B": 18, "C": 18, "D": 18, "E": 20, "F": 16, "G": 16, "H": 32, "I": 16, "J": 24, "K": 24})

    by_month: dict[str, dict] = defaultdict(lambda: {"count": 0, "total": 0, "ramin": 0, "mana": 0, "categories": defaultdict(int)})
    for expense in expenses:
        month_key = to_jalali(expense.expense_date)[:7].replace("/", "-")
        month_data = by_month[month_key]
        month_data["count"] += 1
        month_data["total"] += expense.amount_toman
        month_data[expense.person] += expense.amount_toman
        month_data["categories"][expense.category] += expense.amount_toman

    summary_headers = ["ماه شمسی", "تعداد هزینه", "مجموع (تومان)", "سهم رامین (تومان)", "سهم مانا (تومان)", "بیشترین دسته‌بندی", "مبلغ بیشترین دسته‌بندی (تومان)"]
    summary_sheet.append(["خلاصه ماهانه هزینه‌های خانوار"])
    summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_headers))
    _style_title(summary_sheet["A1"])
    summary_sheet.row_dimensions[1].height = 28
    summary_sheet.append(summary_headers)
    _style_header(summary_sheet[2])
    summary_sheet.freeze_panes = "A3"
    for month_key in sorted(by_month.keys(), reverse=True):
        month_data = by_month[month_key]
        top_category = max(month_data["categories"], key=month_data["categories"].get)
        summary_sheet.append([
            f"{MONTH_NAMES[int(month_key[5:7]) - 1]} {month_key[:4]}",
            month_data["count"], month_data["total"], month_data["ramin"], month_data["mana"],
            category_labels.get(top_category, top_category), month_data["categories"][top_category],
        ])
    if summary_sheet.max_row >= 3:
        _add_table(summary_sheet, f"A2:G{summary_sheet.max_row}", "MonthlySummary")
    for row in summary_sheet.iter_rows(min_row=3, max_row=summary_sheet.max_row):
        for cell in row[2:5]:
            cell.number_format = '#,##0 "تومان"'
        row[6].number_format = '#,##0 "تومان"'
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="right")
    _set_widths(summary_sheet, {"A": 20, "B": 16, "C": 20, "D": 20, "E": 20, "F": 22, "G": 28})

    accounts_sheet = workbook.create_sheet("حساب‌ها")
    account_headers = ["شناسه کاربر", "نام کاربری", "فرد", "وضعیت", "تاریخ ایجاد"]
    accounts_sheet.append(["حساب‌های خانوار — بدون رمز عبور"])
    accounts_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(account_headers))
    _style_title(accounts_sheet["A1"])
    accounts_sheet.row_dimensions[1].height = 28
    accounts_sheet.append(account_headers)
    _style_header(accounts_sheet[2])
    users_by_id = {user.id: user for user in users}
    users_by_id.update({expense.created_by.id: expense.created_by for expense in expenses if getattr(expense, "created_by", None)})
    for account in sorted(users_by_id.values(), key=lambda item: item.id):
        accounts_sheet.append([
            account.id, account.username, PERSON_LABELS.get(account.person, account.person),
            "فعال" if account.is_active else "غیرفعال", _excel_datetime(account.created_at),
        ])
    accounts_sheet.freeze_panes = "A3"
    if accounts_sheet.max_row >= 3:
        _add_table(accounts_sheet, f"A2:E{accounts_sheet.max_row}", "AccountsExport")
    _set_widths(accounts_sheet, {"A": 16, "B": 22, "C": 18, "D": 16, "E": 24})
    for row in accounts_sheet.iter_rows(min_row=3, max_row=accounts_sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="right")

    guide_sheet = workbook.create_sheet("راهنما")
    guide_sheet.append(["راهنمای فایل خروجی"])
    guide_sheet.merge_cells("A1:B1")
    _style_title(guide_sheet["A1"], fill_color=WISE_GREEN)
    guide_sheet.append(["عنوان", "توضیح"])
    _style_header(guide_sheet[2])
    for row in [
        ["منبع", "خروجی مستقیم از دیتابیس خانوار خرج‌نگار"],
        ["تاریخ تولید", "زمان تولید فایل در metadata درخواست ثبت می‌شود."],
        ["مبلغ", "مبالغ به تومان ذخیره شده‌اند؛ ستون هزار تومان برای ورود اولیه نیز وجود دارد."],
        ["تاریخ شمسی", "تاریخ نمایشی جلالی است."],
        ["تاریخ میلادی", "تاریخ استاندارد ذخیره‌شده در دیتابیس است."],
        ["امنیت", "این فایل پس از دانلود روی دستگاه شما قرار می‌گیرد؛ آن را در محل امن نگهداری کنید."],
    ]:
        guide_sheet.append(row)
    guide_sheet.freeze_panes = "A3"
    _add_table(guide_sheet, f"A2:B{guide_sheet.max_row}", "ExportGuide")
    _set_widths(guide_sheet, {"A": 22, "B": 92})
    for row in guide_sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="right")

    for ws in workbook.worksheets:
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
