from fastapi.testclient import TestClient

from app.main import app


def login(client: TestClient) -> str:
    response = client.post(
        "/api/auth/login",
        json={"username": "ramin", "password": "ramin-test-password-123"},
    )
    assert response.status_code == 200
    return client.cookies.get("kharj_csrf")


def test_export_requires_authentication(client):
    response = client.get("/api/export/xlsx")
    assert response.status_code == 401


def test_export_returns_xlsx_download_with_expected_sheets(client):
    login(client)
    response = client.get("/api/export/xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert response.content[:2] == b"PK"
    assert len(response.content) > 1000

    import io
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    assert workbook.sheetnames == ["هزینه‌ها", "خلاصه ماهانه", "حساب‌ها", "راهنما"]
    assert workbook["هزینه‌ها"].max_row >= 2
    assert workbook["خلاصه ماهانه"].max_row >= 2
    assert workbook["حساب‌ها"].max_row >= 2
    assert workbook["راهنما"].max_row >= 2
    workbook.close()
